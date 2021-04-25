import os
from openpyxl import load_workbook
from openpyxl import Workbook

path = r"C:\projects\Pypractice\files\\"
file_list = os.listdir(path)

results = []
for file_name_raw in file_list:
    file_name = path + file_name_raw # 저장경로 + 파일명
    wb = load_workbook(filename=file_name, data_only=True) #이미 존재하는 엑셀파일 열기
    ws = wb.active # 현재 활성화된 worksheet 얻기(파일안에 여러개의 worksheet가 있음)
    result = []
    result.append(file_name_raw)
    result.append(ws['C2'].value)
    result.append(ws['C3'].value)
    result.append(ws['E2'].value)
    result.append(ws['E3'].value)
    results.append(result)
# print(results)

wb = Workbook() #엑셀 파일을 의미,새로 생성
ws = wb.active  #현재 활성화된 workshheet 얻기
for i in results:
    ws.append(i)
wb.save("results.xlsx")